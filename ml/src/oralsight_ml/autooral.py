"""Prepare an audited, training-only Autooral segmentation supplement.

The dataset itself is never copied into the repository. The generated CSV contains
only pseudonymous identifiers, relative controlled-data paths, provenance, and
license terms.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from .release_training import SUPPLEMENTAL_SEGMENTATION_COLUMNS

PROVENANCE_URL = "https://github.com/wurenkai/HF-UNet-and-Autooral-dataset"
SOURCE_DATASET = "Autooral-SciRep-2024"
LICENSE_TERMS = "academic_research_noncommercial"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_workbook(path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            'Autooral preparation requires the optional "research" dependencies.'
        ) from exc
    return load_workbook(path, read_only=True, data_only=True)


def _normalized_image_number(value: object) -> str:
    if value is None:
        raise ValueError("Image Number is blank.")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text.isdigit():
        raise ValueError(f"Image Number is not numeric: {text!r}")
    return text.zfill(4)


def _prepare_rows(dataset_root: Path, archive_sha256: str) -> list[dict[str, str]]:
    workbook_path = dataset_root / "Training Set disease classification.xlsx"
    if not workbook_path.is_file():
        raise ValueError("Autooral training workbook was not found.")
    workbook = _load_workbook(workbook_path)
    values = list(workbook.active.iter_rows(values_only=True))
    if not values:
        raise ValueError("Autooral training workbook is empty.")
    header = [str(value).strip() if value is not None else "" for value in values[0]]
    try:
        patient_index = header.index("Patient ID")
        image_index = header.index("Image Number")
    except ValueError as exc:
        raise ValueError("Autooral workbook lacks Patient ID or Image Number.") from exc

    rows: list[dict[str, str]] = []
    seen_images: set[str] = set()
    for row_number, value_row in enumerate(values[1:], start=2):
        patient_id = str(value_row[patient_index]).strip()
        if not patient_id or patient_id == "None":
            raise ValueError(f"Patient ID is blank on workbook row {row_number}.")
        image_number = _normalized_image_number(value_row[image_index])
        if image_number in seen_images:
            raise ValueError(f"Duplicate Image Number in workbook: {image_number}")
        seen_images.add(image_number)
        image_path = f"Train/data_train/{image_number}.png"
        mask_path = f"Train/mask_train/{image_number}.png"
        if not (dataset_root / image_path).is_file():
            raise ValueError(f"Training image is missing: {image_path}")
        if not (dataset_root / mask_path).is_file():
            raise ValueError(f"Training mask is missing: {mask_path}")
        rows.append(
            {
                "sample_id": f"autooral-train-{image_number}",
                "patient_id": f"autooral-train-patient-{patient_id}",
                "split": "train",
                "image_path": image_path,
                "mask_path": mask_path,
                "source_dataset": SOURCE_DATASET,
                "device_family": "mixed-clinical-camera",
                "license_status": "approved",
                "audit_status": "approved",
                "consent_scope": "research_training",
                "license_terms": LICENSE_TERMS,
                "provenance_url": PROVENANCE_URL,
                "archive_sha256": archive_sha256,
            }
        )

    image_files = {path.stem for path in (dataset_root / "Train" / "data_train").glob("*.png")}
    mask_files = {path.stem for path in (dataset_root / "Train" / "mask_train").glob("*.png")}
    if image_files != seen_images or mask_files != seen_images:
        raise ValueError(
            "Autooral workbook, training images, and training masks do not have identical IDs."
        )
    return rows


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset-root", type=Path, required=True)
    parser.add_argument("--output-manifest", type=Path, required=True)
    parser.add_argument("--archive-sha256", required=True)
    parser.add_argument("--acknowledge-academic-only-license", action="store_true")
    parser.add_argument("--acknowledge-audited-data", action="store_true")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    if not args.acknowledge_academic_only_license:
        print(
            "Preparation refused: acknowledge the academic, non-commercial license.",
            file=sys.stderr,
        )
        return 2
    if not args.acknowledge_audited_data:
        print("Preparation refused: acknowledge the local data audit.", file=sys.stderr)
        return 2
    if not args.dataset_root.is_dir():
        print("Preparation refused: dataset root does not exist.", file=sys.stderr)
        return 2
    archive_sha256 = args.archive_sha256.strip().lower()
    if len(archive_sha256) != 64 or any(
        character not in "0123456789abcdef" for character in archive_sha256
    ):
        print("Preparation refused: archive SHA-256 is invalid.", file=sys.stderr)
        return 2
    if args.output_manifest.exists():
        print("Preparation refused: output manifest already exists.", file=sys.stderr)
        return 2
    try:
        rows = _prepare_rows(args.dataset_root, archive_sha256)
        args.output_manifest.parent.mkdir(parents=True, exist_ok=True)
        with args.output_manifest.open("x", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=SUPPLEMENTAL_SEGMENTATION_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"Preparation failed safely: {exc}", file=sys.stderr)
        return 2
    summary = {
        "manifest": args.output_manifest.name,
        "manifest_sha256": _sha256(args.output_manifest),
        "sample_count": len(rows),
        "patient_count": len({row["patient_id"] for row in rows}),
        "source_dataset": SOURCE_DATASET,
        "license_terms": LICENSE_TERMS,
        "training_only": True,
    }
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
